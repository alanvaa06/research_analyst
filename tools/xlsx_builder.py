"""Deterministic xlsx style/scaffold engine for the research_analyst plugin.

Doctrine: format is CODE, not model judgement. model-standards builds every
workbook THROUGH this module (never raw openpyxl styling), and /model-check
runs ``audit_format`` (checks F) against any workbook, plugin-built or not.

Empirical basis: CFI workbook corpus (3-Statement Model Complete, AMZN
Advanced case, Valuation Model, template library) extracted 2026-08-30.

Console output policy: ASCII only ([ok]/[x], no unicode symbols).

Usage as CLI:
    python tools/xlsx_builder.py audit <path.xlsx>     -> run checks F, exit 1 on failure
    python tools/xlsx_builder.py demo  <path.xlsx>     -> build a skeleton (self-test)
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from enum import Enum, auto
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Contract constants (the ONLY allowed values; checks F enforce the whitelists)
# ---------------------------------------------------------------------------

BUILDER_STAMP_KEY = "research_analyst_builder"
BUILDER_STAMP_VALUE = "xlsx_builder v1"

FONT_NAME = "Aptos Narrow"


class Color(str, Enum):
    """Palette (ARGB). Identidad propia del plugin (2026-08-31): azul oscuro +
    azul claro. Los valores CFI legacy se conservan como referencia historica
    pero YA NO estan en la whitelist."""

    DARK_BLUE = "FF1F4E79"   # barra de marca + bandas de seccion (texto blanco)
    LIGHT_BLUE = "FFBDD7EE"  # sub-secciones + acento / tab color
    NAVY = "FF132E57"        # legacy CFI (fuera de whitelist)
    ORANGE = "FFED942D"      # legacy CFI (fuera de whitelist)
    TEAL = "FF1E8496"        # legacy CFI (fuera de whitelist)
    INPUT_BLUE = "FF0000FF"  # analyst input font
    LINK_GREEN = "FF00CC00"  # cross-sheet link font
    WARN_RED = "FFFF0000"    # error font
    WHITE = "FFFFFFFF"
    BLACK = "FF000000"
    INPUT_FILL = "FFFFF2CC"  # light yellow input shading
    SCENARIO_FILL = "FFF2F2F2"  # light gray scenario areas


class NumFmt(str, Enum):
    """Number-format whitelist (literal CFI strings)."""

    GENERAL = "General"
    NUM = '_-* #,##0_-;\\(#,##0\\)_-;_-* "-"_-;_-@_-'   # thousands, (neg), dash zero
    NUM_RED = "#,##0_);[Red](#,##0);-"                   # red negatives variant
    PCT1 = "0.0%"
    PCT2 = "0.00%"
    DEC2 = "0.00"
    MULT = "0.0\\x"                                       # 12.3x
    USD = '"$"#,##0_);\\("$"#,##0\\)'
    USD_CENTS = '"$"#,##0.00_);\\("$"#,##0.00\\)'
    YEAR_A = '0"A"'                                       # 2025A
    YEAR_E = '0"E"'                                       # 2026E
    DATE = "mm-dd-yy"
    HIDDEN = ";;;"


class CellRole(Enum):
    """Semantic cell roles -> font color mapping (traceability contract)."""

    LABEL = auto()      # black text
    INPUT = auto()      # blue font + yellow fill (analyst assumption)
    OBSERVED = auto()   # blue font, NO fill (historical, cited in comment)
    FORMULA = auto()    # black font
    LINK = auto()       # green font (pulls from another sheet)
    WARN = auto()       # red font


_ROLE_FONT_COLOR: dict[CellRole, Color] = {
    CellRole.LABEL: Color.BLACK,
    CellRole.INPUT: Color.INPUT_BLUE,
    CellRole.OBSERVED: Color.INPUT_BLUE,
    CellRole.FORMULA: Color.BLACK,
    CellRole.LINK: Color.LINK_GREEN,
    CellRole.WARN: Color.WARN_RED,
}

_THIN = Side(style="thin")
_DOUBLE = Side(style="double")

# Sheets that must carry frozen panes (data grids). Cover/Checks/Summary exempt.
FROZEN_SHEET_PREFIXES = ("Operating", "Annual", "Model", "Assumptions",
                         "Macro", "IS", "BS", "CF", "Ratios", "Schedules",
                         "Rev_Reconcile", "Val_", "Quarterly")

# Ratios completeness contract (check F13): these labels must exist in the
# Ratios section — build_ratios() writes exactly these, so builder output
# passes by construction and a lazy hand-built Ratios fails.
REQUIRED_RATIO_LABELS = (
    "Margen neto", "Rotacion de activos", "Apalancamiento", "ROE DuPont 3",
    "Carga fiscal", "Carga de interes", "Margen EBIT", "ROE DuPont 5",
    "NOPAT", "Capital invertido", "ROIC", "Economic profit",
    "Margen bruto", "Margen operativo", "Razon corriente", "Quick ratio",
    "Deuda / EBITDA aprox", "Cobertura de intereses",
    "DSO", "DIO", "DPO", "CCC",
    "DFL", "CFO / NI", "Accruals",
)

JUNK_SHEET_NAMES = ("Hoja1", "Hoja2", "Sheet1", "Sheet2", "Hoja 1", "Sheet 1")


@dataclass(frozen=True)
class PeriodHeader:
    """Year header spec: e.g. 2019..2031, actuals through 2025."""

    first_year: int
    last_year: int
    last_actual_year: int


# Rebrandable DECORATIVE slots (brand/DESIGN.md). Semantic colors — input blue,
# link green, warn red, input/scenario fills — are the traceability contract
# and are NEVER brandable.
BRAND_SLOTS = ("brand_primary", "brand_section", "brand_accent")


def load_brand(path: str) -> dict[str, str]:
    """Parse brand/DESIGN.md lines like ``brand_primary: #132E57`` -> ARGB.

    Deterministic: only the three BRAND_SLOTS are read; anything else in the
    file is prose for humans. Missing file or missing slot -> CFI default.
    """
    import re
    try:
        with open(path, encoding="utf-8") as fh:
            text = fh.read()
    except OSError:
        return {}
    out: dict[str, str] = {}
    for slot in BRAND_SLOTS:
        match = re.search(rf"{slot}\s*[:=]\s*#?([0-9A-Fa-f]{{6}})", text)
        if match:
            out[slot] = "FF" + match.group(1).upper()
    return out


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


class ModelStyler:
    """Owns a Workbook and applies the format contract. One instance per model."""

    def __init__(self, units_label: str = "USD millones salvo indicado",
                 brand: Optional[dict[str, str]] = None) -> None:
        self.wb: Workbook = Workbook()
        self.units_label = units_label
        b = brand or {}
        self.color_primary: str = b.get("brand_primary", Color.DARK_BLUE.value)
        self.color_section: str = b.get("brand_section", Color.DARK_BLUE.value)
        self.color_accent: str = b.get("brand_accent", Color.LIGHT_BLUE.value)
        default = self.wb.active
        if default is not None:
            self.wb.remove(default)
        self.wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)
        self._stamp()

    # -- workbook level -----------------------------------------------------

    def _stamp(self) -> None:
        """Custom doc property proving builder provenance (check F10)."""
        self.wb.properties.keywords = f"{BUILDER_STAMP_KEY}={BUILDER_STAMP_VALUE}"

    def set_periodicity(self, mode: str) -> None:
        """Stamp the model's periodicity (issuer-profile) into the workbook so
        check F14 can verify quarter columns without external context."""
        self.wb.properties.keywords = (
            f"{self.wb.properties.keywords or ''};periodicity={mode}")

    def define_constant(self, name: str, sheet: str, coord: str) -> None:
        """Named range for a labeled constant (e.g. DAYS_YEAR) — kills hardcodes."""
        from openpyxl.workbook.defined_name import DefinedName
        ref = f"'{sheet}'!${coord[0]}${coord[1:]}"
        self.wb.defined_names.add(DefinedName(name, attr_text=ref))

    def save(self, path: str) -> None:
        self.wb.save(path)

    # -- sheet level --------------------------------------------------------

    def new_sheet(self, name: str, freeze: Optional[str] = "C4",
                  tab_color: Optional[str] = None) -> Worksheet:
        """tab_color: ARGB string; use self.color_accent for branded tabs."""
        ws = self.wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        if freeze:
            ws.freeze_panes = freeze
        if tab_color is not None:
            ws.sheet_properties.tabColor = tab_color
        return ws

    def brand_bar(self, ws: Worksheet, title: str, last_col: int = 18) -> None:
        """Rows 1-2: brand bar + sheet title + units note. Row 3 reserved for checks."""
        for col in range(1, last_col + 1):
            for row in (1, 2):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill("solid", fgColor=self.color_primary)
        c = ws.cell(row=1, column=1, value="(c) research_analyst - todos los supuestos son del analista")
        c.font = Font(name=FONT_NAME, size=8, color=Color.WHITE.value)
        t = ws.cell(row=2, column=1, value=title)
        t.font = Font(name=FONT_NAME, size=16, bold=True, color=Color.WHITE.value)
        u = ws.cell(row=2, column=4, value=f"({self.units_label})")
        u.font = Font(name=FONT_NAME, size=11, color=Color.WHITE.value)

    def period_header(self, ws: Worksheet, row: int, first_col: int,
                      spec: PeriodHeader) -> None:
        """Year row with A/E suffix formats (2025A / 2026E)."""
        col = first_col
        for year in range(spec.first_year, spec.last_year + 1):
            cell = ws.cell(row=row, column=col, value=year)
            fmt = NumFmt.YEAR_A if year <= spec.last_actual_year else NumFmt.YEAR_E
            cell.number_format = fmt.value
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")
            col += 1

    def check_row(self, ws: Worksheet, row: int, first_col: int, n_cols: int,
                  formula_template: str) -> None:
        """Per-column check row near the top (frozen visible). Template uses {col}."""
        label = ws.cell(row=row, column=2, value="Balance Sheet Check")
        label.font = Font(name=FONT_NAME, size=11, bold=True)
        self._nav_mark(ws, row)
        for i in range(n_cols):
            col_letter = get_column_letter(first_col + i)
            cell = ws.cell(row=row, column=first_col + i,
                           value=formula_template.format(col=col_letter))
            cell.font = Font(name=FONT_NAME, size=11, color=Color.WARN_RED.value)
            cell.alignment = Alignment(horizontal="center")

    # -- row/cell level -----------------------------------------------------

    def _nav_mark(self, ws: Worksheet, row: int,
                  color: str = Color.BLACK.value) -> None:
        """Navigation pattern: column A carries ONLY an 'x' on each header/
        sub-header row, so Ctrl+arrow on column A jumps section to section.
        Labels live in column B."""
        c = ws.cell(row=row, column=1, value="x")
        c.font = Font(name=FONT_NAME, size=8, color=color)

    def section_header(self, ws: Worksheet, row: int, title: str,
                       last_col: int = 18) -> None:
        """Banda azul oscuro, texto BLANCO bold 16 — una por seccion."""
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                "solid", fgColor=self.color_section)
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name=FONT_NAME, size=16, bold=True,
                      color=Color.WHITE.value)
        self._nav_mark(ws, row, color=Color.WHITE.value)

    def subsection(self, ws: Worksheet, row: int, title: str,
                   last_col: int = 18) -> None:
        """Banda azul claro, texto negro bold 14 — sub-jerarquia."""
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                "solid", fgColor=self.color_accent)
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name=FONT_NAME, size=14, bold=True)
        self._nav_mark(ws, row)

    def schedule_block_header(self, ws: Worksheet, row: int, name: str) -> None:
        """Header of one 'Sch: <name>' block inside the single Schedules sheet."""
        self.subsection(ws, row, f"Sch: {name}")

    def set_cell(self, ws: Worksheet, coord: str, value: object, role: CellRole,
                 numfmt: NumFmt = NumFmt.NUM, size: int = 11,
                 bold: bool = False) -> None:
        cell = ws[coord]
        cell.value = value
        cell.font = Font(name=FONT_NAME, size=size, bold=bold,
                         color=_ROLE_FONT_COLOR[role].value)
        cell.number_format = numfmt.value
        if role is CellRole.INPUT:
            cell.fill = PatternFill("solid", fgColor=Color.INPUT_FILL.value)

    def series_row(self, ws: Worksheet, row: int, label: str, first_col: int,
                   hist_values: list[object], forecast_values: list[object],
                   numfmt: NumFmt = NumFmt.NUM,
                   hist_role: CellRole = CellRole.FORMULA,
                   forecast_role: CellRole = CellRole.INPUT) -> None:
        """ONE continuous row across the whole horizon (contract F11).

        A series is never split into 'historical' and 'forecast' rows: the same
        line carries computed/observed history (black or blue-observed) and then
        forecast cells (blue input on yellow, or driver formula). The role
        switches at the boundary column; every horizon column gets a value.
        """
        lab = ws.cell(row=row, column=2, value=label)
        lab.font = Font(name=FONT_NAME, size=11, color=Color.BLACK.value)
        col = first_col
        for value in hist_values:
            self.set_cell(ws, f"{get_column_letter(col)}{row}", value,
                          hist_role, numfmt)
            col += 1
        for value in forecast_values:
            self.set_cell(ws, f"{get_column_letter(col)}{row}", value,
                          forecast_role, numfmt)
            col += 1

    def subtotal_border(self, ws: Worksheet, row: int, first_col: int,
                        n_cols: int) -> None:
        for i in range(n_cols):
            ws.cell(row=row, column=first_col + i).border = Border(top=_THIN)

    def total_border(self, ws: Worksheet, row: int, first_col: int,
                     n_cols: int) -> None:
        for i in range(n_cols):
            ws.cell(row=row, column=first_col + i).border = Border(
                top=_THIN, bottom=_DOUBLE)

    def group_rows(self, ws: Worksheet, start: int, end: int,
                   hidden: bool = False) -> None:
        """Outline level 1 so sections collapse to summary view."""
        for r in range(start, end + 1):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = hidden

    def label_col_width(self, ws: Worksheet, width: float = 42.0) -> None:
        """Column A = narrow navigation column ('x' markers); B = labels."""
        ws.column_dimensions["A"].width = 2.5
        ws.column_dimensions["B"].width = width

    def quarter_header(self, ws: Worksheet, row: int, first_col: int,
                       quarters: list[str]) -> int:
        """Quarter labels ('1Q2026E' / '3Q2025A'). Returns the next free column."""
        col = first_col
        for label in quarters:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")
            col += 1
        return col

    def interleaved_header(self, ws: Worksheet, row: int, first_col: int,
                           first_year: int, last_year: int,
                           quarterly_from_year: int, last_actual_year: int,
                           last_actual_quarter: int) -> dict[tuple[int, str], int]:
        """Quarterly-native header: per fiscal year, 4 quarter columns then the
        FY aggregate column; years before ``quarterly_from_year`` get FY only.

        Returns {(year, '1Q'|'2Q'|'3Q'|'4Q'|'FY'): column} so the build wires
        FY = aggregate-of-quarters formulas (C8 estructural) and populate knows
        where each period lives.
        """
        colmap: dict[tuple[int, str], int] = {}
        col = first_col
        for year in range(first_year, last_year + 1):
            if year >= quarterly_from_year:
                for q in (1, 2, 3, 4):
                    is_actual = (year < last_actual_year or
                                 (year == last_actual_year and q <= last_actual_quarter))
                    label = f"{q}Q{year}{'A' if is_actual else 'E'}"
                    cell = ws.cell(row=row, column=col, value=label)
                    cell.font = Font(name=FONT_NAME, size=11, bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    colmap[(year, f"{q}Q")] = col
                    col += 1
            fy = ws.cell(row=row, column=col, value=year)
            fy.number_format = (NumFmt.YEAR_A if year <= last_actual_year
                                else NumFmt.YEAR_E).value
            fy.font = Font(name=FONT_NAME, size=11, bold=True)
            fy.alignment = Alignment(horizontal="center")
            colmap[(year, "FY")] = col
            col += 1
        return colmap

    # -- Ratios section (deterministic; fixes the "lazy ratios" failure) -----

    def build_ratios(self, ws: Worksheet, start_row: int, first_col: int,
                     n_cols: int, ref: dict[str, str],
                     wacc_ref: Optional[str] = None) -> tuple[int, list[str]]:
        """Write the FULL Ratios section (blocks A-G of ratios-analytics.md).

        ``ref`` maps canon line -> absolute row reference WITHOUT column, e.g.
        {"rev": "Model!{c}27", ...} where "{c}" is replaced per period column
        and "{p}" by the previous column. Required canons: rev, cogs, gross,
        ebit, ebt, ni, interest, tax, ta, equity, cash, ar, inv, ap, ca, cl,
        debt, re, cfo, da. Missing canons skip their rows (returned in the
        skipped list) — but check F13 fails if the section is incomplete, so
        a skip is visible, never silent.
        """
        skipped: list[str] = []
        r = start_row
        # Auto-generar referencias de periodo previo: el caller pasa canons
        # planos ({c}); aqui se derivan los <canon>_p ({p}) que usan las
        # plantillas — evita el footgun de claves manuales.
        ref = dict(ref)
        for k, v in list(ref.items()):
            if not k.endswith("_p") and isinstance(v, str) and "{c}" in v:
                ref.setdefault(k + "_p", v.replace("{c}", "{p}"))

        def row_out(label: str, template: str, fmt: NumFmt,
                    needs: tuple[str, ...]) -> None:
            nonlocal r
            if any(k not in ref for k in needs):
                skipped.append(label)
                return
            ws.cell(row=r, column=2, value=label).font = Font(
                name=FONT_NAME, size=11)
            for i in range(1, n_cols):  # first period column has no prior year
                col = get_column_letter(first_col + i)
                prev = get_column_letter(first_col + i - 1)
                parts = {k: v.replace("{c}", col).replace("{p}", prev)
                         for k, v in ref.items()}
                formula = template.format(**parts)
                self.set_cell(ws, f"{col}{r}", formula, CellRole.FORMULA, fmt)
            r += 1

        def header(title: str) -> None:
            nonlocal r
            self.section_header(ws, r, title)
            r += 1

        AVG = "AVERAGE({p_ref},{c_ref})"

        def avg(canon: str) -> dict[str, str]:
            return {}

        header("DuPont")
        row_out("Margen neto (NI/Ventas)", '=IF({rev}=0,"",{ni}/{rev})', NumFmt.PCT1, ("ni", "rev"))
        row_out("Rotacion de activos (Ventas/Activos prom.)",
                '=IF(AVERAGE({ta_p},{ta})=0,"",{rev}/AVERAGE({ta_p},{ta}))', NumFmt.DEC2, ("rev", "ta", "ta_p"))
        row_out("Apalancamiento (Activos/Capital prom.)",
                '=IF(AVERAGE({equity_p},{equity})=0,"",AVERAGE({ta_p},{ta})/AVERAGE({equity_p},{equity}))', NumFmt.DEC2, ("ta", "ta_p", "equity", "equity_p"))
        row_out("ROE DuPont 3", '=IF(AVERAGE({equity_p},{equity})=0,"",{ni}/AVERAGE({equity_p},{equity}))', NumFmt.PCT1, ("ni", "equity", "equity_p"))
        row_out("Carga fiscal (NI/EBT)", '=IF({ebt}=0,"",{ni}/{ebt})', NumFmt.PCT1, ("ni", "ebt"))
        row_out("Carga de interes (EBT/EBIT)", '=IF({ebit}=0,"",{ebt}/{ebit})', NumFmt.PCT1, ("ebt", "ebit"))
        row_out("Margen EBIT (EBIT/Ventas)", '=IF({rev}=0,"",{ebit}/{rev})', NumFmt.PCT1, ("ebit", "rev"))
        row_out("ROE DuPont 5 (producto)",
                '=IF(OR({ebt}=0,{ebit}=0,{rev}=0,AVERAGE({equity_p},{equity})=0),"",'
                '{ni}/{ebt}*{ebt}/{ebit}*{ebit}/{rev}*{rev}/AVERAGE({ta_p},{ta})'
                '*AVERAGE({ta_p},{ta})/AVERAGE({equity_p},{equity}))', NumFmt.PCT1,
                ("ni", "ebt", "ebit", "rev", "ta", "ta_p", "equity", "equity_p"))
        r += 1

        header("ROIC y economic profit")
        row_out("Tasa efectiva (tax/EBT)", '=IF({ebt}=0,"",{tax}/{ebt})', NumFmt.PCT1, ("tax", "ebt"))
        row_out("NOPAT (EBIT x (1-t))", '=IF({ebt}=0,"",{ebit}*(1-{tax}/{ebt}))', NumFmt.NUM, ("ebit", "tax", "ebt"))
        row_out("Capital invertido (deuda+capital-caja)", "={debt}+{equity}-{cash}", NumFmt.NUM, ("debt", "equity", "cash"))
        row_out("ROIC", '=IF(({debt}+{equity}-{cash})=0,"",IF({ebt}=0,"",{ebit}*(1-{tax}/{ebt})/({debt}+{equity}-{cash})))', NumFmt.PCT1, ("debt", "equity", "cash", "ebit", "tax", "ebt"))
        if wacc_ref:
            row_out(f"Economic profit (spread vs WACC {wacc_ref})",
                    '=IF({ebt}=0,"",({ebit}*(1-{tax}/{ebt})/MAX(1,{debt}+{equity}-{cash})-' + wacc_ref + ')*({debt}+{equity}-{cash}))',
                    NumFmt.NUM, ("ebit", "tax", "ebt", "debt", "equity", "cash"))
        else:
            skipped.append("Economic profit (sin wacc_ref)")
        r += 1

        header("Rentabilidad, liquidez y solvencia")
        row_out("Margen bruto", '=IF({rev}=0,"",{gross}/{rev})', NumFmt.PCT1, ("gross", "rev"))
        row_out("Margen operativo", '=IF({rev}=0,"",{ebit}/{rev})', NumFmt.PCT1, ("ebit", "rev"))
        row_out("Razon corriente", '=IF({cl}=0,"",{ca}/{cl})', NumFmt.DEC2, ("ca", "cl"))
        row_out("Quick ratio", '=IF({cl}=0,"",({ca}-{inv})/{cl})', NumFmt.DEC2, ("ca", "cl", "inv"))
        row_out("Deuda / EBITDA aprox (EBIT+D&A)", '=IF(({ebit}+{da})=0,"",{debt}/({ebit}+{da}))', NumFmt.DEC2, ("debt", "ebit", "da"))
        row_out("Cobertura de intereses (EBIT/interes)", '=IF({interest}=0,"n/a",{ebit}/ABS({interest}))', NumFmt.DEC2, ("ebit", "interest"))
        r += 1

        header("Ciclo de conversion de efectivo")
        row_out("DSO (dias)", '=IF({rev}=0,"",AVERAGE({ar_p},{ar})/{rev}*DAYS_YEAR)', NumFmt.DEC2, ("ar", "ar_p", "rev"))
        row_out("DIO (dias)", '=IF({cogs}=0,"",AVERAGE({inv_p},{inv})/{cogs}*DAYS_YEAR)', NumFmt.DEC2, ("inv", "inv_p", "cogs"))
        row_out("DPO (dias)", '=IF({cogs}=0,"",AVERAGE({ap_p},{ap})/{cogs}*DAYS_YEAR)', NumFmt.DEC2, ("ap", "ap_p", "cogs"))
        row_out("CCC (DSO+DIO-DPO)",
                '=IF({cogs}=0,"",AVERAGE({ar_p},{ar})/{rev}*DAYS_YEAR+AVERAGE({inv_p},{inv})/{cogs}*DAYS_YEAR-AVERAGE({ap_p},{ap})/{cogs}*DAYS_YEAR)',
                NumFmt.DEC2, ("ar", "ar_p", "inv", "inv_p", "ap", "ap_p", "rev", "cogs"))
        r += 1

        header("Apalancamiento operativo y calidad")
        row_out("DFL (EBIT/(EBIT-interes))", '=IF(({ebit}-ABS({interest}))=0,"",{ebit}/({ebit}-ABS({interest})))', NumFmt.DEC2, ("ebit", "interest"))
        row_out("CFO / NI (calidad de utilidades)", '=IF({ni}=0,"",{cfo}/{ni})', NumFmt.DEC2, ("cfo", "ni"))
        row_out("Accruals proxy (NI-CFO)/Activos prom.",
                '=IF(AVERAGE({ta_p},{ta})=0,"",({ni}-{cfo})/AVERAGE({ta_p},{ta}))', NumFmt.PCT1, ("ni", "cfo", "ta", "ta_p"))
        return r, skipped


# ---------------------------------------------------------------------------
# Checks F — format audit (deterministic, any workbook)
# ---------------------------------------------------------------------------

_ALLOWED_FONT_COLORS = {c.value for c in (
    Color.INPUT_BLUE, Color.LINK_GREEN, Color.WARN_RED, Color.WHITE, Color.BLACK)}
_ALLOWED_FONT_COLORS.add("FF333333")  # near-black tolerated
_ALLOWED_FILLS = {c.value for c in (
    Color.DARK_BLUE, Color.LIGHT_BLUE, Color.INPUT_FILL, Color.SCENARIO_FILL)}
_ALLOWED_NUMFMTS = {f.value for f in NumFmt}

_MAX_SCAN_ROWS = 400
_MAX_SCAN_COLS = 40


@dataclass(frozen=True)
class Finding:
    check: str
    ok: bool
    detail: str


def _scan_fonts_fills_formats(ws: Worksheet) -> tuple[set[str], set[str], set[str], set[str]]:
    font_names: set[str] = set()
    font_colors: set[str] = set()
    fills: set[str] = set()
    numfmts: set[str] = set()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, _MAX_SCAN_ROWS),
                            max_col=min(ws.max_column, _MAX_SCAN_COLS)):
        for cell in row:
            if cell.value is None:
                continue
            font = cell.font
            if font is not None and font.name:
                font_names.add(str(font.name))
                rgb = getattr(font.color, "rgb", None) if font.color else None
                if isinstance(rgb, str):
                    font_colors.add(rgb)
            if cell.fill is not None and cell.fill.patternType == "solid":
                rgb = getattr(cell.fill.fgColor, "rgb", None)
                if isinstance(rgb, str) and rgb != "00000000":
                    fills.add(rgb)
            numfmts.add(cell.number_format)
    return font_names, font_colors, fills, numfmts


def _series_continuity_violations(ws: Worksheet) -> list[str]:
    """F11: rows that look like input series must span the WHOLE horizon.

    Period columns = columns whose header cell uses the 0"A"/0"E" year formats.
    Any row with >= 3 input-filled cells among period columns is a series row;
    a series row with empty period cells means history/forecast got split or
    history was left unpopulated (the AAPL smoke-test failure pattern).
    """
    period_cols: list[int] = []
    header_row: Optional[int] = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8),
                            max_col=min(ws.max_column, _MAX_SCAN_COLS)):
        cols = [c.column for c in row
                if c.number_format in (NumFmt.YEAR_A.value, NumFmt.YEAR_E.value)
                and c.value is not None]
        if len(cols) >= 4:
            period_cols = cols
            header_row = row[0].row
            break
    if not period_cols or header_row is None:
        return []
    violations: list[str] = []
    for r in range(header_row + 1, min(ws.max_row, _MAX_SCAN_ROWS) + 1):
        filled_inputs = 0
        empties = 0
        for col in period_cols:
            cell = ws.cell(row=r, column=col)
            is_input_fill = (cell.fill is not None
                             and cell.fill.patternType == "solid"
                             and getattr(cell.fill.fgColor, "rgb", None)
                             == Color.INPUT_FILL.value)
            if cell.value is not None and is_input_fill:
                filled_inputs += 1
            if cell.value is None:
                empties += 1
        if filled_inputs >= 3 and empties > 0:
            violations.append(f"{ws.title}!fila {r} ({empties} celdas vacias)")
    return violations


def audit_format(path: str, brand: Optional[dict[str, str]] = None) -> list[Finding]:
    """Run checks F1-F11 on a workbook. Pure read; returns findings.

    ``brand``: output of load_brand(brand/DESIGN.md) — its values extend the
    fill whitelist (F4) so a branded model audits green with its own DESIGN.md.
    """
    allowed_fills = _ALLOWED_FILLS | set((brand or {}).values())
    wb = load_workbook(path, data_only=False)
    findings: list[Finding] = []
    visible = [wb[n] for n in wb.sheetnames if wb[n].sheet_state == "visible"]

    # F1 gridlines off everywhere
    bad = [ws.title for ws in visible if ws.sheet_view.showGridLines in (True, None)]
    findings.append(Finding("F1 gridlines off", not bad, ", ".join(bad) or "todas ok"))

    all_fonts: set[str] = set()
    all_colors: set[str] = set()
    all_fills: set[str] = set()
    all_fmts: set[str] = set()
    for ws in visible:
        fn, fc, fl, nf = _scan_fonts_fills_formats(ws)
        all_fonts |= fn
        all_colors |= fc
        all_fills |= fl
        all_fmts |= nf

    # F2 single standard font family
    alien_fonts = sorted(all_fonts - {FONT_NAME})
    findings.append(Finding("F2 fuente estandar", not alien_fonts,
                            ", ".join(alien_fonts) or FONT_NAME))
    # F3 font colors within whitelist
    alien_colors = sorted(all_colors - _ALLOWED_FONT_COLORS)
    findings.append(Finding("F3 colores de fuente", not alien_colors,
                            ", ".join(alien_colors) or "whitelist ok"))
    # F4 fills within palette (+ brand slots if DESIGN.md provided)
    alien_fills = sorted(all_fills - allowed_fills)
    findings.append(Finding("F4 paleta de fills", not alien_fills,
                            ", ".join(alien_fills) or "whitelist ok"))
    # F5 number formats within whitelist
    alien_fmts = sorted(all_fmts - _ALLOWED_NUMFMTS)
    findings.append(Finding("F5 formatos numericos", not alien_fmts,
                            "; ".join(alien_fmts[:6]) or "whitelist ok"))
    # F6 frozen panes on data sheets
    no_freeze = [ws.title for ws in visible
                 if ws.title.startswith(FROZEN_SHEET_PREFIXES) and not ws.freeze_panes]
    findings.append(Finding("F6 freeze panes", not no_freeze,
                            ", ".join(no_freeze) or "ok"))
    # F7 outline grouping: en la tab Model (secciones apiladas) o en Schedules
    host = next((ws for ws in visible if ws.title in ("Model", "Schedules")), None)
    hosts = [ws for ws in visible
             if ws.title in ("Operating", "Annual", "Model", "Schedules")]
    if not hosts:
        findings.append(Finding("F7 outline por seccion", False,
                                "sin hojas Operating/Annual/Model/Schedules"))
    else:
        # Por SECCION (marcador 'x' en col A + bold >=13 en col B): cada
        # seccion con contenido debe tener filas agrupadas — agrupar solo
        # algunas secciones (bug del smoke 2026-08-31) FALLA.
        ungrouped: list[str] = []
        total_grouped = 0
        for host in hosts:
            total_grouped += sum(1 for d in host.row_dimensions.values()
                                 if d.outlineLevel)
            headers: list[int] = []
            for r in range(1, min(host.max_row, _MAX_SCAN_ROWS) + 1):
                a = host.cell(row=r, column=1).value
                b = host.cell(row=r, column=2)
                if (isinstance(a, str) and a.strip().lower() == "x"
                        and b.font is not None and b.font.bold
                        and (b.font.size or 0) >= 13):
                    headers.append(r)
            for i, hr in enumerate(headers):
                end = (headers[i + 1] - 1 if i + 1 < len(headers)
                       else min(host.max_row, _MAX_SCAN_ROWS))
                content = [r for r in range(hr + 1, end + 1)
                           if any(host.cell(row=r, column=c).value is not None
                                  for c in range(2, min(host.max_column, 30) + 1))]
                if not content:
                    continue
                grouped = sum(1 for r in content
                              if r in host.row_dimensions
                              and host.row_dimensions[r].outlineLevel)
                if grouped == 0:
                    label = host.cell(row=hr, column=2).value
                    ungrouped.append(f"{host.title}!fila {hr} ({str(label)[:25]})")
        ok = total_grouped > 0 and not ungrouped
        detail = (f"{total_grouped} filas agrupadas"
                  + (f"; secciones SIN outline: {', '.join(ungrouped[:5])}"
                     if ungrouped else ""))
        findings.append(Finding("F7 outline por seccion", ok, detail))
    # F8 no junk sheets, no Sch_* sheets
    junk = [n for n in wb.sheetnames if n in JUNK_SHEET_NAMES or n.startswith("Sch_")]
    findings.append(Finding("F8 sin hojas basura/Sch_*", not junk,
                            ", ".join(junk) or "ok"))
    # F9 A/E period formats present somewhere
    has_ae = any(f in all_fmts for f in (NumFmt.YEAR_A.value, NumFmt.YEAR_E.value))
    findings.append(Finding("F9 headers de periodo A/E", has_ae,
                            "presentes" if has_ae else "sin formato 0\"A\"/0\"E\""))
    # F10 builder stamp
    kw = wb.properties.keywords or ""
    stamped = BUILDER_STAMP_KEY in kw
    findings.append(Finding("F10 sello del builder", stamped,
                            kw if stamped else "sin sello (no construido por xlsx_builder)"))
    # F11 series continuity: input-series rows span the whole horizon
    all_violations: list[str] = []
    for ws in visible:
        all_violations.extend(_series_continuity_violations(ws))
    findings.append(Finding("F11 continuidad de series", not all_violations,
                            "; ".join(all_violations[:8]) or
                            "filas de input completas en todo el horizonte"))
    # F12 series partidas: fila etiquetada "forecast"/"historico" con solo una
    # mitad del horizonte poblada = la serie se partio en dos filas (violacion
    # de "una serie = una fila"; el patron exacto del smoke AAPL)
    split_hits: list[str] = []
    for ws in visible:
        split_hits.extend(_split_series_violations(ws))
    findings.append(Finding("F12 sin series partidas", not split_hits,
                            "; ".join(split_hits[:8]) or
                            "ninguna fila hist/forecast partida"))
    # F13 completitud + UNICIDAD de Ratios: el set completo presente, y cada
    # razon UNA sola vez por hoja — un label duplicado delata secciones
    # "Ratios historico" / "Ratios forecast" partidas (bug del smoke #3);
    # la serie completa vive en UNA fila.
    counts: dict[tuple[str, str], int] = {}
    for ws in visible:
        if ws.title not in ("Operating", "Annual", "Model", "Ratios"):
            continue
        for r in range(1, min(ws.max_row, _MAX_SCAN_ROWS) + 1):
            a_cell = ws.cell(row=r, column=1)
            b_cell = ws.cell(row=r, column=2)
            # Saltar encabezados de seccion/sub-seccion (mismo criterio que
            # F7: bold >=13) — sus titulos contienen nombres de razones
            # ("ROIC y economic profit") y NO son filas de razon.
            def _is_header(c) -> bool:
                return (c.font is not None and c.font.bold
                        and (c.font.size or 0) >= 13)
            if _is_header(a_cell) or _is_header(b_cell):
                continue
            label = b_cell.value
            if not isinstance(label, str) or not label.strip():
                label = a_cell.value
            if not isinstance(label, str):
                continue
            low = label.strip().lower()
            if low == "x":
                continue
            # Comparacion por PREFIJO: "CCC (DSO+DIO-DPO)" cuenta solo para
            # CCC, no para DSO/DIO/DPO (que aparecen como substring).
            for req in REQUIRED_RATIO_LABELS:
                if low.startswith(req.lower()):
                    key_ = (ws.title, req)
                    counts[key_] = counts.get(key_, 0) + 1
    labels_found = {label for (_, label) in counts}
    missing_ratios = [x for x in REQUIRED_RATIO_LABELS if x not in labels_found]
    dup_ratios = [f"{sheet}:{label}" for (sheet, label), n in counts.items()
                  if n >= 2]
    ok13 = not missing_ratios and not dup_ratios
    if missing_ratios:
        detail13 = ("faltan: " + ", ".join(missing_ratios[:8])
                    + (" ..." if len(missing_ratios) > 8 else ""))
    elif dup_ratios:
        detail13 = ("secciones de Ratios PARTIDAS (label duplicado): "
                    + ", ".join(dup_ratios[:5]))
    else:
        detail13 = f"{len(REQUIRED_RATIO_LABELS)} razones presentes, sin duplicados"
    findings.append(Finding("F13 Ratios completa y unica", ok13, detail13))
    # F14 columnas trimestrales estimadas: si el sello dice periodicidad con
    # trimestres, el header debe traer columnas #Q20yyE (el contrato 1a que el
    # rebuild v3 se salto). Sin sello de periodicidad: n/a (modelo externo).
    import re as _re
    m = _re.search(r"periodicity=([a-z_]+)", kw)
    if m and m.group(1) in ("annual_plus_quarterly", "quarterly"):
        # (i) Operating (o Model legacy): el modelo SE CONSTRUYE sobre
        # trimestres — >=4 A y >=4 E en el header.
        q_actual = q_est = 0
        for ws in visible:
            if ws.title not in ("Operating", "Model"):
                continue
            for row in ws.iter_rows(min_row=1, max_row=8,
                                    max_col=min(ws.max_column, 200)):
                for c in row:
                    if isinstance(c.value, str):
                        v = c.value.strip()
                        if _re.fullmatch(r"[1-4]Q20\d\dA", v):
                            q_actual += 1
                        elif _re.fullmatch(r"[1-4]Q20\d\dE", v):
                            q_est += 1
        problems: list[str] = []
        if q_actual < 4 or q_est < 4:
            problems.append(f"Operating: {q_actual} trimestres A / {q_est} E (min 4 y 4)")
        # (ii) Annual: solo anios FY en header (cero '#Q') y CERO inputs —
        # los agregados jamas se teclean (C8 estructural).
        annual = next((ws for ws in visible if ws.title == "Annual"), None)
        if annual is not None:
            q_in_annual = 0
            inputs_in_annual = 0
            for row in annual.iter_rows(min_row=1, max_row=8,
                                        max_col=min(annual.max_column, 60)):
                for c in row:
                    if isinstance(c.value, str) and _re.fullmatch(
                            r"[1-4]Q20\d\d[AE]", c.value.strip()):
                        q_in_annual += 1
            for row in annual.iter_rows(max_row=min(annual.max_row, _MAX_SCAN_ROWS),
                                        max_col=min(annual.max_column, 60)):
                for c in row:
                    if (c.fill is not None and c.fill.patternType == "solid"
                            and getattr(c.fill.fgColor, "rgb", None)
                            == Color.INPUT_FILL.value and c.value is not None):
                        inputs_in_annual += 1
            if q_in_annual:
                problems.append(f"Annual: {q_in_annual} columnas #Q (debe ser FY-solo)")
            if inputs_in_annual:
                problems.append(f"Annual: {inputs_in_annual} celdas de INPUT (agregados no se teclean)")
        findings.append(Finding(
            "F14 modelo trimestral-nativo", not problems,
            "; ".join(problems) if problems else
            f"Operating: {q_actual} A / {q_est} E; Annual limpio"))
    else:
        findings.append(Finding("F14 modelo trimestral-nativo", True,
                                "n/a (sin sello de periodicidad trimestral)"))
    return findings


def _split_series_violations(ws: Worksheet) -> list[str]:
    """F12: label contiene 'forecast' u 'historico' y solo su mitad esta llena."""
    period_cols_a: list[int] = []
    period_cols_e: list[int] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8),
                            max_col=min(ws.max_column, _MAX_SCAN_COLS)):
        for c in row:
            if c.value is None:
                continue
            if c.number_format == NumFmt.YEAR_A.value:
                period_cols_a.append(c.column)
            elif c.number_format == NumFmt.YEAR_E.value:
                period_cols_e.append(c.column)
        if period_cols_a and period_cols_e:
            break
    if not period_cols_a or not period_cols_e:
        return []
    hits: list[str] = []
    for r in range(1, min(ws.max_row, _MAX_SCAN_ROWS) + 1):
        label = ws.cell(row=r, column=1).value
        if not isinstance(label, str) or label.strip().lower() == "x":
            label = ws.cell(row=r, column=2).value
        if not isinstance(label, str):
            continue
        low = label.lower()
        is_fc = "forecast" in low
        is_hist = "historico" in low or "histórico" in low
        if not (is_fc or is_hist):
            continue
        a_vals = sum(1 for c in period_cols_a if ws.cell(row=r, column=c).value is not None)
        e_vals = sum(1 for c in period_cols_e if ws.cell(row=r, column=c).value is not None)
        if is_fc and e_vals >= 2 and a_vals == 0:
            hits.append(f"{ws.title}!fila {r} (solo forecast)")
        if is_hist and a_vals >= 2 and e_vals == 0:
            hits.append(f"{ws.title}!fila {r} (solo historico)")
    return hits


def _print_report(findings: Iterable[Finding]) -> int:
    failures = 0
    for f in findings:
        mark = "[ok]" if f.ok else "[x]"
        if not f.ok:
            failures += 1
        print(f"{mark} {f.check}: {f.detail}")
    print(f"Resumen F: {sum(1 for f in findings if f.ok)} ok, {failures} fallas")
    return 1 if failures else 0


def _demo(path: str) -> None:
    """Self-test skeleton: proves the builder passes its own audit."""
    styler = ModelStyler()
    spec = PeriodHeader(first_year=2019, last_year=2031, last_actual_year=2025)
    for name in ("Cover", "Checks", "Assumptions", "Macro", "IS", "BS", "CF",
                 "Ratios", "Schedules", "Rev_Reconcile", "Val_DCF", "Val_Comps",
                 "Sensitivity", "Summary"):
        freeze = "C4" if name.startswith(FROZEN_SHEET_PREFIXES) else None
        ws = styler.new_sheet(name, freeze=freeze)
        styler.brand_bar(ws, name)
        styler.label_col_width(ws)
        if name.startswith(FROZEN_SHEET_PREFIXES):
            styler.period_header(ws, 3, 3, spec)
    sched = styler.wb["Schedules"]
    row = 5
    for block in ("PPE", "Debt", "WC"):
        styler.schedule_block_header(sched, row, block)
        styler.group_rows(sched, row + 1, row + 4)
        styler.subtotal_border(sched, row + 4, 3, 13)
        row += 6
    assum = styler.wb["Assumptions"]
    styler.subsection(assum, 5, "Drivers (demo)")
    styler.series_row(assum, 6, "Crecimiento unidades (%)", 3,
                      hist_values=[0.05] * 7, forecast_values=[0.04] * 6,
                      numfmt=NumFmt.PCT1)
    ratios = styler.wb["Ratios"]
    for i, label in enumerate(REQUIRED_RATIO_LABELS):
        ratios.cell(row=5 + i, column=2, value=label).font = Font(
            name=FONT_NAME, size=11)
    styler.save(path)
    print(f"[ok] demo escrito: {path}")


def main(argv: list[str]) -> int:
    if argv[1:2] == ["audit"] and len(argv) in (3, 4):
        brand = load_brand(argv[3]) if len(argv) == 4 else None
        return _print_report(audit_format(argv[2], brand=brand))
    if len(argv) == 3 and argv[1] == "demo":
        _demo(argv[2])
        return _print_report(audit_format(argv[2]))
    print("uso: python xlsx_builder.py audit <path.xlsx> [brand/DESIGN.md] | demo <path.xlsx>")
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
